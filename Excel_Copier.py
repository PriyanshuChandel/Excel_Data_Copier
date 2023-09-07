from os import makedirs
from openpyxl import load_workbook
from warnings import filterwarnings
from threading import Thread
from time import time
from os.path import join, dirname, splitext, basename, exists
from tkinter import Tk, filedialog, Label, Button, Entry, Toplevel, Canvas, Checkbutton, IntVar
from datetime import datetime
from tkinter.ttk import Progressbar, Style, Combobox
from tkinter.scrolledtext import ScrolledText
from random import randint
from xlwings import Book, App
from sys import argv


class ExApp:
    if not exists('log'):
        makedirs('log')
    logFilCount = 1
    fileHandler = None
    maxLinesPerLogFile = 49999
    currentLineCountLog = 0
    iconFile = join(dirname(__file__), 'Excel_Mac_23559.ico')
    aboutIcon = join(dirname(__file__), 'info.ico')

    sourceExcelFile = ''
    targetExcelFile = ''

    copySelectedIndex = []
    copyCheckBoxVars = []
    copyCheckboxesText = []
    copyCheckboxes = {}
    copySelectedIndices = []
    copyColors = {}
    copySelectedValues = ''

    pasteColors = {}
    pasteSelectedIndex = []
    pasteCheckBoxVars = []
    pasteCheckboxesText = []
    pasteCheckboxes = {}
    pasteSelectedIndices = []
    pasteSelectedValues = ''

    sourceExcelSheetComboValues = ['Select WorkSheet']
    targetExcelSheetComboValues = ['Select WorkSheet']
    sourceReferenceColumnComboValues = ['Select Reference Column']
    targetReferenceColumnComboValues = ['Select Reference Column']

    def __init__(self):
        self.window = Tk()
        self.window.config(bg='lemon chiffon')
        self.window.title('ExcelBridge v1.2')
        self.window.geometry('450x525')
        self.window.resizable(False, False)
        self.window.iconbitmap(self.iconFile)
        self.mainLabel = Label(self.window, text='Excel Data Transfer', font=('Arial', 15, 'bold'), fg='blue',
                               bg='lemon chiffon')
        self.mainLabel.place(x=230, y=15, anchor='center')

        self.sourceEntryCanvas = Canvas(self.window, highlightthickness=3, highlightbackground="black",
                                        relief='solid', height=100, width=435, bg='gray56')
        self.sourceEntryCanvas.place(x=5, y=30)
        self.sourceExcelLabel = Label(self.sourceEntryCanvas, text='Source Excel:', font=('Arial', 9, 'bold italic'),
                                      bg='gray56')
        self.sourceExcelLabel.place(x=5, y=10)
        self.sourceExcelEntry = Entry(self.sourceEntryCanvas, bd=4, width=47, bg='white', state='disabled')
        self.sourceExcelEntry.place(x=120, y=8)
        self.sourceExcelFileDialogBtn = Button(self.sourceEntryCanvas, text='...', bg='green', fg='white',
                                               font=('Arial', 8), cursor='hand2',
                                               command=self.sourceExcelFileDialogFunc)
        self.sourceExcelFileDialogBtn.place(x=418, y=8)
        self.sourceExcelSheetLabel = Label(self.sourceEntryCanvas, text='Source Excel Sheet:',
                                           font=('Arial', 8, 'bold italic'), bg='gray56')
        self.sourceExcelSheetLabel.place(x=5, y=40)
        self.sourceExcelSheetCombo = Combobox(self.sourceEntryCanvas, values=self.sourceExcelSheetComboValues,
                                              width=47, state='disabled', cursor='arrow', validate="key")
        self.sourceExcelSheetCombo.place(x=120, y=38)
        self.sourceExcelSheetCombo.current(0)
        self.sourceReferenceColumnLabel = Label(self.sourceEntryCanvas, text='Source Ref. column:',
                                                font=('Arial', 8, 'bold italic'), bg='gray56')
        self.sourceReferenceColumnLabel.place(x=5, y=70)
        self.sourceReferenceColumnCombo = Combobox(self.sourceEntryCanvas,
                                                   values=self.sourceReferenceColumnComboValues, width=47,
                                                   state='disabled', cursor='arrow', validate="key")
        self.sourceReferenceColumnCombo.place(x=120, y=68)
        self.sourceReferenceColumnCombo.current(0)

        self.targetEntryCanvas = Canvas(self.window, highlightthickness=3, highlightbackground="yellow",
                                        relief='solid', height=100, width=435, bg='gray56')
        self.targetEntryCanvas.place(x=5, y=140)
        self.targetExcelLabel = Label(self.targetEntryCanvas, text='Target Excel:', font=('Arial', 8, 'bold italic'),
                                      bg='grey56')
        self.targetExcelLabel.place(x=5, y=10)
        self.targetExcelEntry = Entry(self.targetEntryCanvas, bd=4, width=47, bg='white', state='disabled')
        self.targetExcelEntry.place(x=120, y=8)
        self.targetExcelFileDialogBtn = Button(self.targetEntryCanvas, text='...', bg='light grey', fg='white',
                                               font=('Arial', 8), state='disabled', cursor='arrow',
                                               command=self.targetExcelFileDialogFunc)
        self.targetExcelFileDialogBtn.place(x=418, y=8)
        self.targetExcelSheetLabel = Label(self.targetEntryCanvas, text='Target Excel Sheet:',
                                           font=('Arial', 8, 'bold italic'), bg='grey56')
        self.targetExcelSheetLabel.place(x=5, y=40)

        self.targetExcelSheetCombo = Combobox(self.targetEntryCanvas, values=self.targetExcelSheetComboValues, width=47,
                                              state='disabled', cursor='arrow', validate="key")
        self.targetExcelSheetCombo.place(x=120, y=38)
        self.targetExcelSheetCombo.current(0)
        self.targetReferenceColumnLabel = Label(self.targetEntryCanvas, text='Target Ref. column:',
                                                font=('Arial', 8, 'bold italic'), bg='grey56')
        self.targetReferenceColumnLabel.place(x=5, y=70)

        self.targetReferenceColumnCombo = Combobox(self.targetEntryCanvas, values=self.targetReferenceColumnComboValues,
                                                   width=47, state='disabled', cursor='arrow', validate="key")
        self.targetReferenceColumnCombo.place(x=120, y=68)
        self.targetReferenceColumnCombo.current(0)

        self.copyPasteCanvas = Canvas(self.window, highlightthickness=3, highlightbackground="light green",
                                      relief='solid', height=265, width=435, bg='gray56')
        self.copyPasteCanvas.place(x=5, y=250)

        self.sourceCopyColumnsLabel = Label(self.copyPasteCanvas, text='Copy Columns', font=('Arial', 8, 'bold italic'),
                                            bg='grey56')
        self.sourceCopyColumnsLabel.place(x=100, y=5)
        self.sourceCopyColumnsScrolledText = ScrolledText(self.copyPasteCanvas, bd=4, width=12, bg='white', height=5)
        self.sourceCopyColumnsScrolledText.place(x=100, y=25)
        self.targetPasteColumnsLabel = Label(self.copyPasteCanvas, text='Paste Columns:',
                                             font=('Arial', 8, 'bold italic'), bg='grey56')
        self.targetPasteColumnsLabel.place(x=245, y=5)
        self.targetPasteColumnsScrolledText = ScrolledText(self.copyPasteCanvas, bd=4, width=12, bg='white', height=5)
        self.targetPasteColumnsScrolledText.place(x=245, y=25)

        self.submitBtn = Button(self.copyPasteCanvas, text='Submit', bg='light grey', fg='white',
                                font=('Arial', 12, 'bold'), cursor='arrow',
                                state='disabled', command=self.matchCopyPaste)
        self.submitBtn.place(x=155, y=120)
        self.resetBtn = Button(self.copyPasteCanvas, text='Reset', bg='light grey', fg='white',
                               font=('Arial', 12, 'bold'), state='disabled', cursor='arrow',
                               command=self.resetBtnFunc)
        self.resetBtn.place(x=245, y=120)
        self.progress = Progressbar(self.copyPasteCanvas, length=430, mode="determinate",
                                    style="Custom.Horizontal.TProgressbar")
        self.progress.place(x=5, y=160)
        self.progressStyle = Style()
        self.progressStyle.theme_use('default')
        self.progressStyle.configure("Custom.Horizontal.TProgressbar", thickness=20, troughcolor='#E0E0E0',
                                     background='#FFFF00',
                                     troughrelief='flat', relief='flat', text='0 %')
        self.progressStyle.layout('Custom.Horizontal.TProgressbar', [('Horizontal.Progressbar.trough',
                                                                      {'children': [('Horizontal.Progressbar.pbar',
                                                                                     {'side': 'left', 'sticky': 'ns'})],
                                                                       'sticky': 'nswe'}),
                                                                     ('Horizontal.Progressbar.label', {'sticky': ''})])
        self.warnLabel = Label(self.copyPasteCanvas, text='Note:It does not supports Data_Validation_Rules',
                               font=('Arial', 8, 'bold italic'), bg='grey56')
        self.warnLabel.place(x=5, y=245)
        self.messageLabel = Label(self.copyPasteCanvas, justify='left', wraplength=420, font=('Arial', 8, 'bold'),
                                  bg='grey56')
        self.messageLabel.place(x=5, y=195)
        self.aboutLabel = Label(self.copyPasteCanvas, text="About", bg='grey56', fg="blue", cursor="hand2")
        self.aboutLabel.place(x=395, y=245)

    def aboutWindow(self):
        aboutWin = Toplevel(self.window)
        aboutWin.grab_set()
        aboutWin.geometry('285x90')
        aboutWin.resizable(False, False)
        aboutWin.title('About')
        aboutWin.iconbitmap(self.aboutIcon)
        aboutWinLabel = Label(aboutWin,
                              text=f'Version - 1.2\nDeveloped by Priyanshu\nFor any improvement please reach on '
                                   f'below email\nEmail : chandelpriyanshu8@outlook.com\nMobile : '
                                   f'+91-8285775109 '
                                   f'', font=('Helvetica', 9)).place(x=1, y=6)

    def updateProgress(self, newVal, totalVal):
        resultVal = round((newVal / totalVal) * 98, 2)
        self.progress['value'] = resultVal
        self.progressStyle.configure("Custom.Horizontal.TProgressbar", text=f"{resultVal}%")
        self.window.update()

    def resetBtnFunc(self):
        self.writeLog('info', 'Resetting...')
        self.sourceExcelEntry.config(state='normal')
        self.sourceExcelEntry.delete(0, 'end')
        self.sourceExcelEntry.config(state='disabled')
        self.sourceExcelSheetCombo.config(state='normal')
        self.sourceExcelSheetCombo.delete(0, 'end')
        self.sourceExcelSheetCombo.set('Select WorkSheet')
        self.sourceExcelSheetCombo.config(state='disabled', cursor='arrow')
        self.sourceReferenceColumnCombo.config(state='normal')
        self.sourceReferenceColumnCombo.delete(0, 'end')
        self.sourceReferenceColumnCombo.set('Select Reference Column')
        self.sourceReferenceColumnCombo.config(state='disabled', cursor='arrow')
        self.sourceExcelFileDialogBtn.config(state='normal', bg='green', cursor='hand2')
        self.sourceExcelFile = ''
        self.copySelectedIndices.clear()
        self.copyCheckBoxVars.clear()
        self.copyCheckboxes.clear()
        self.copySelectedIndex.clear()
        self.copyCheckboxesText.clear()
        self.copyColors.clear()
        self.sourceCopyColumnsScrolledText.delete(1.0, 'end')
        self.targetExcelEntry.config(state='normal')
        self.targetExcelEntry.delete(0, 'end')
        self.targetExcelEntry.config(state='disabled')
        self.targetExcelSheetCombo.config(state='normal')
        self.targetExcelSheetCombo.delete(0, 'end')
        self.targetExcelSheetCombo.set('Select WorkSheet')
        self.targetExcelSheetCombo.config(state='disabled', cursor='arrow')
        self.targetReferenceColumnCombo.config(state='normal')
        self.targetReferenceColumnCombo.delete(0, 'end')
        self.targetReferenceColumnCombo.set('Select Reference Column')
        self.targetReferenceColumnCombo.config(state='disabled', cursor='arrow')
        self.targetExcelFileDialogBtn.config(state='disabled', bg='light grey', cursor='arrow')
        self.pasteSelectedIndices.clear()
        self.pasteCheckBoxVars.clear()
        self.pasteCheckboxes.clear()
        self.pasteSelectedIndex.clear()
        self.pasteCheckboxesText.clear()
        self.pasteColors.clear()
        self.targetPasteColumnsScrolledText.delete(1.0, 'end')
        self.submitBtn.config(state='disabled', bg='light grey', cursor='arrow')
        self.resetBtn.config(state='disabled', bg='light grey', cursor='arrow')
        self.targetExcelFile = ''
        self.messageLabel.config(text='')
        self.aboutLabel.config(stat='normal', cursor='hand2')
        self.aboutLabel.bind("<Button-1>", lambda event: self.aboutWindow())
        self.writeLog('info', 'Reset done..')

    def writeLog(self, messageType, message):
        timeStamp = datetime.now().strftime('%Y_%m_%d_%H_%M_%S')
        formattedFileCount = f"{self.logFilCount:05d}"
        newFileName = f'logs_{timeStamp}_({formattedFileCount}).txt'
        if self.fileHandler is None or self.currentLineCountLog >= self.maxLinesPerLogFile:
            if self.fileHandler is not None:
                self.fileHandler.write(f"End of file, Next file: {newFileName}")
                self.fileHandler.flush()
                self.fileHandler.close()
            self.fileHandler = open(f"log/{newFileName}", 'a')
            self.logFilCount += 1
            self.currentLineCountLog = 0
        self.fileHandler.write(f'{datetime.now().replace(microsecond=0)} [{messageType}] {message}\n')
        self.currentLineCountLog += 1
        self.fileHandler.flush()

    def getColumnLetter(self, index):
        try:
            if index <= 0:
                self.writeLog('error', '[ValueError][Column index must be greater than 0]')
            result = ''
            while index > 0:
                index, remainder = divmod(index - 1, 26)
                result = chr(65 + remainder) + result
            return result
        except Exception as e:
            self.writeLog('error', 'Error while converting index into excel column alphabet')
            self.writeLog('debug', f'{e}')
            self.messageLabel.config(text='error, check logs!')

    def excelColumnIndex(self, columnRef):
        try:
            base = ord('A') - 1
            index = 0
            for char in columnRef:
                index = index * 26 + (ord(char) - base)
            return index
        except Exception as e:
            self.writeLog('error', 'Error while converting Excel Columns Reference into its index')
            self.writeLog('debug', f'{e}')
            self.messageLabel.config(text='error, check logs!')

    def randomColor(self):
        try:
            return f'#{randint(0, 0xFFFFFF):06x}'
        except Exception as e:
            self.writeLog('error', 'Error while getting random color')
            self.writeLog('debug', f'{e}')
            self.messageLabel.config(text='error, check logs!')

    def disableCheckboxes(self, checkboxes, exceptCheckbox=None):
        try:
            for checkbox, var in checkboxes.items():
                if checkbox != exceptCheckbox:
                    checkbox.config(state='disabled')
        except Exception as e:
            self.writeLog('error', 'Error while disabling checkboxes')
            self.writeLog('debug', f'{e}')
            self.messageLabel.config(text='error, check logs!')

    def enableCheckboxes(self, checkboxes):
        try:
            for checkbox, var in checkboxes.items():
                if var.get() == 0:
                    checkbox.config(state='normal')
                else:
                    checkbox.config(state='disabled')
        except Exception as e:
            self.writeLog('error', 'Error while enabling checkboxes')
            self.writeLog('debug', f'{e}')
            self.messageLabel.config(text='error, check logs!')

    def assignColor(self, checkBoxVar, checkboxes, colors):
        color = self.randomColor()
        for checkbox, var in checkboxes.items():
            if var == checkBoxVar:
                checkbox.configure(bg=color)
                checkbox.config(state='disabled')
                colors[checkbox] = color

    def sourceExcelFileDialogFunc(self):
        def innerDialogThread():
            self.aboutLabel.config(stat='disabled', cursor='arrow')
            self.aboutLabel.unbind("<Button-1>")
            self.messageLabel.config(text='')
            self.progress.config(value=0)
            self.progressStyle.configure("Custom.Horizontal.TProgressbar", text='0 %')
            self.sourceExcelSheetComboValues.clear()
            self.sourceExcelSheetComboValues.append('Select WorkSheet')
            self.sourceExcelFile = filedialog.askopenfilename(filetypes=(("Excel files", "*.xlsx"),))
            self.sourceExcelEntry.config(state='normal')
            self.sourceExcelEntry.delete(0, 'end')
            self.sourceExcelEntry.insert(0, self.sourceExcelFile)
            self.sourceExcelEntry.config(state='readonly')
            filterwarnings("ignore", category=UserWarning)
            sourceExcelEnt = self.sourceExcelEntry.get()
            if len(sourceExcelEnt) > 0:
                self.writeLog('info', f'[{self.sourceExcelFile}] is selected as source excel file')
                self.sourceExcelFileDialogBtn.config(state='disabled', bg='light grey', cursor='arrow')
                try:
                    sourceWorkBook = load_workbook(sourceExcelEnt)
                    for sourceExcelSheet in sourceWorkBook.sheetnames:
                        self.sourceExcelSheetComboValues.append(sourceExcelSheet)
                    self.sourceExcelSheetCombo.config(state='normal', cursor='hand2')
                    self.sourceExcelSheetCombo.config(values=self.sourceExcelSheetComboValues)
                    self.resetBtn.config(state='normal', bg='red', cursor='hand2')
                    sourceWorkBook.close()
                except PermissionError as permError:
                    self.writeLog('error', f'[Permission Error] [Permission Denied] [{self.sourceExcelFile}] is opened '
                                           f'by another application, close it first')
                    self.writeLog('debug', f'{permError}')
                    self.messageLabel.config(text='Close excel file, in use by another app')
                    self.sourceExcelFileDialogBtn.config(state='normal', bg='green', cursor='hand2')
                except IndexError as indError:
                    self.writeLog('error', f'[Index Error] Unable to access the sheets of excel file file '
                                           f'[{self.sourceExcelFile}]. Possibly because excel file is not correctly '
                                           f'saved before')
                    self.writeLog('debug', f'{indError}')
                    self.messageLabel.config(text='Error, Try again with different excel file')
                    self.sourceExcelFileDialogBtn.config(state='normal', bg='green', cursor='hand2')
                except Exception as e:
                    self.writeLog('error', f'[Unhandled Error] Unable to access the excel file '
                                           f'[{self.sourceExcelFile}] check --> Format, Permission, Sensitivity '
                                           f'setting etc.')
                    self.writeLog('debug', f'{e}')
                    self.messageLabel.config(text='Error, check file format permissions, sensitivity setting etc.')
                    self.sourceExcelFileDialogBtn.config(state='normal', bg='green', cursor='hand2')
            else:
                self.writeLog('warning', 'No source excel file selected')
                self.messageLabel.config(text='Select source Excel')

        threadSourceExcelFileDialog = Thread(target=innerDialogThread)
        threadSourceExcelFileDialog.start()

    def enableSourceReferenceColumnCombo(self):
        def innerSourceThread():
            self.sourceCopyColumnsScrolledText.delete(1.0, 'end')
            sourceComboSelection = self.sourceExcelSheetCombo.get()
            if not sourceComboSelection == "Select WorkSheet":
                self.sourceReferenceColumnCombo.config(state='disabled', cursor='arrow')
                self.sourceExcelSheetCombo.config(state='disabled', cursor='arrow')
                self.resetBtn.config(state='disabled', bg='light grey', cursor='arrow')
                self.sourceReferenceColumnComboValues.clear()
                self.sourceReferenceColumnComboValues.append('Select Reference Column')
                sourceWorkBook = load_workbook(self.sourceExcelFile)
                for sourceColumnIndex in range(1, sourceWorkBook[sourceComboSelection].max_column + 1):
                    sourceColumnLetter = self.getColumnLetter(sourceColumnIndex)
                    self.sourceReferenceColumnComboValues.append(sourceColumnLetter)
                self.sourceExcelSheetCombo.config(state='normal', cursor='hand2')
                self.sourceReferenceColumnCombo.config(state='normal', cursor='hand2')
                self.sourceReferenceColumnCombo.config(values=self.sourceReferenceColumnComboValues)
                del self.sourceReferenceColumnComboValues[0]
                sourceWorkBook.close()
            else:
                self.sourceReferenceColumnCombo.config(state='disabled', cursor='arrow')
            self.resetBtn.config(state='normal', bg='red', cursor='hand2')

        threadEnableSourceReferenceColumnCombo = Thread(target=innerSourceThread)
        threadEnableSourceReferenceColumnCombo.start()

    def enableTargetExcelDialogBtn(self, ):
        sourceExcelSheetSelected = self.sourceExcelSheetCombo.get()
        sourceRefCol = self.sourceReferenceColumnCombo.get()
        if sourceExcelSheetSelected != 'Select WorkSheet' and sourceRefCol != 'Select Reference Column':
            self.sourceExcelSheetCombo.config(state='disabled', cursor='arrow')
            self.sourceReferenceColumnCombo.config(state='disabled', cursor='arrow')
            self.sourceReferenceColumnComboValues.remove(sourceRefCol)
            for index, columnsText in enumerate(self.sourceReferenceColumnComboValues):
                var = IntVar()
                self.copyCheckBoxVars.append(var)
                copyCheckbox = Checkbutton(self.sourceCopyColumnsScrolledText, variable=var, text=columnsText,
                                           bg='light grey', state='disabled', cursor="arrow",
                                           command=lambda e=index: self.copyCheckboxClicked(e))
                copyCheckbox.pack()
                self.copyCheckboxesText.append(columnsText)
                self.copyCheckboxes[copyCheckbox] = var
                self.sourceCopyColumnsScrolledText.window_create('end', window=copyCheckbox)
                self.sourceCopyColumnsScrolledText.config(bd=4, width=12, bg='light grey', height=5)
                var.trace_add('write', self.enablePasteCheckBtnSubmitBtn)
            self.targetExcelFileDialogBtn.config(state='normal', bg='green', cursor='hand2')
            sourceExcelFileName = basename(self.sourceExcelFile)
            self.writeLog('info', f'[{sourceExcelSheetSelected}] is selected as source work sheet of the excel '
                                  f'file [{sourceExcelFileName}]')
            self.writeLog('info', f'[{sourceRefCol}] is selected as source reference column from the excel work '
                                  f'sheet [{sourceExcelSheetSelected}]')
        else:
            self.targetExcelFileDialogBtn.config(state='disabled', bg='light grey', cursor='arrow')

    def targetExcelFileDialogFunc(self):
        def innerTargetDialogThread():
            self.messageLabel.config(text='')
            self.progress.config(value=0)
            self.resetBtn.config(state='disabled', bg='light grey', cursor='arrow')
            self.progressStyle.configure("Custom.Horizontal.TProgressbar", text='0 %')
            self.targetExcelSheetComboValues.clear()
            self.targetExcelSheetComboValues.append('Select WorkSheet')
            self.targetExcelFile = filedialog.askopenfilename(filetypes=(("Excel files", "*.xlsx"),))
            self.targetExcelEntry.config(state='normal')
            self.targetExcelEntry.delete(0, 'end')
            self.targetExcelEntry.insert(0, self.targetExcelFile)
            self.targetExcelEntry.config(state='readonly')
            targetExcelEnt = self.targetExcelEntry.get()
            if len(targetExcelEnt) > 0:
                if self.sourceExcelFile == self.targetExcelFile:
                    self.writeLog('info', 'Same source and target excel file selected. Choose different file')
                    self.messageLabel.config(text='Choose different Excel file; target and source cannot be same')
                else:
                    self.writeLog('info', f'[{self.targetExcelFile}] is selected as target excel file')
                    self.targetExcelFileDialogBtn.config(state='disabled', bg='light grey', cursor='arrow')
                    try:
                        targetWorkBook = load_workbook(targetExcelEnt)
                        for targetExcelSheet in targetWorkBook.sheetnames:
                            self.targetExcelSheetComboValues.append(targetExcelSheet)
                        targetWorkBook.close()
                        self.targetExcelSheetCombo.config(state='normal', cursor='hand2')
                        self.targetExcelSheetCombo.config(values=self.targetExcelSheetComboValues)
                    except PermissionError as permError:
                        self.writeLog('error', f'[Permission Error] [Permission Denied] [{self.targetExcelFile}] is '
                                               f'opened by another application, close it first')
                        self.writeLog('debug', f'{permError}')
                        self.messageLabel.config(text='Close excel file, in use by another app')
                        self.targetExcelFileDialogBtn.config(state='normal', bg='green', cursor='hand2')
                    except IndexError as indError:
                        self.writeLog('error', f'[Index Error] Unable to access the sheets of excel '
                                               f'file [{self.targetExcelFile}]. Possibly because excel file is not '
                                               f'correctly saved before')
                        self.writeLog('debug', f'{indError}')
                        self.messageLabel.config(text='Error, Try again with different excel file')
                        self.targetExcelFileDialogBtn.config(state='normal', bg='green', cursor='hand2')
                    except Exception as e:
                        self.writeLog('error', f'[Unhandled Error] Unable to access the excel file '
                                               f'[{self.targetExcelFile}] check --> Format, Permission, Sensitivity '
                                               f'setting etc.')
                        self.writeLog('debug', f'{e}')
                        self.messageLabel.config(text='Error, check file format, permissions, sensitivity setting etc.')
                        self.targetExcelFileDialogBtn.config(state='normal', bg='green', cursor='hand2')
            else:
                self.writeLog('warning', 'No target excel file selected')
                self.messageLabel.config(text='Select target Excel')
            self.resetBtn.config(state='normal', bg='red', cursor='hand2')

        threadTargetExcelFileDialog = Thread(target=innerTargetDialogThread)
        threadTargetExcelFileDialog.start()

    def enableTargetReferenceColumnCombo(self):
        def innerTargetThread():
            self.targetPasteColumnsScrolledText.delete(1.0, 'end')
            targetComboSelection = self.targetExcelSheetCombo.get()
            if not targetComboSelection == "Select WorkSheet":
                self.targetReferenceColumnCombo.config(state='disabled', cursor='arrow')
                self.targetExcelSheetCombo.config(state='disabled', cursor='arrow')
                self.resetBtn.config(state='disabled', bg='light grey', cursor='arrow')
                self.targetReferenceColumnComboValues.clear()
                self.targetReferenceColumnComboValues.append('Select Reference Column')
                targetWorkBook = load_workbook(self.targetExcelFile)
                for targetColumnIndex in range(1, targetWorkBook[targetComboSelection].max_column + 1):
                    targetColumnLetter = self.getColumnLetter(targetColumnIndex)
                    self.targetReferenceColumnComboValues.append(targetColumnLetter)
                self.targetReferenceColumnCombo.config(state='normal', cursor='hand2')
                self.targetExcelSheetCombo.config(state='normal', cursor='hand2')
                self.targetReferenceColumnCombo.config(values=self.targetReferenceColumnComboValues)
                del self.targetReferenceColumnComboValues[0]
                targetWorkBook.close()
            else:
                self.targetReferenceColumnCombo.config(state='disabled', cursor='arrow')
            self.resetBtn.config(state='normal', bg='red', cursor='hand2')

        threadEnableTargetReferenceColumnCombo = Thread(target=innerTargetThread)
        threadEnableTargetReferenceColumnCombo.start()

    def copyCheckboxClicked(self, index):
        if self.copyCheckBoxVars[index].get() == 1:
            self.assignColor(self.copyCheckBoxVars[index], self.copyCheckboxes, self.copyColors)
            self.copySelectedIndices.append(index)
            self.disableCheckboxes(self.copyCheckboxes)
            self.enableCheckboxes(self.pasteCheckboxes)
        elif self.copyCheckBoxVars[index].get() == 0:
            self.enableCheckboxes(self.copyCheckboxes)
            self.copySelectedIndices.remove(index)
        self.copySelectedValues = [self.copyCheckboxesText[i] for i in self.copySelectedIndices]
        self.copySelectedIndex.clear()
        for item in self.copySelectedValues:
            self.copySelectedIndex.append(self.excelColumnIndex(item))

    def pasteCheckboxClicked(self, index):
        pasteCheckVar = self.pasteCheckBoxVars[index].get()
        pasteCheckKeys = self.pasteCheckboxes.keys()
        if pasteCheckVar == 1:
            self.pasteSelectedIndices.append(index)
            self.disableCheckboxes(self.pasteCheckboxes, exceptCheckbox=list(pasteCheckKeys)[index])
            if len(pasteCheckKeys) != len(self.pasteSelectedIndices):
                self.enableCheckboxes(self.copyCheckboxes)
            if len(self.copyColors) > 0:
                color = list(self.copyColors.values())[-1]
                list(pasteCheckKeys)[index].configure(bg=color)
                list(pasteCheckKeys)[index].config(state='disabled')
                self.pasteColors[list(pasteCheckKeys)[index]] = color
            else:
                self.assignColor(self.pasteCheckBoxVars[index], self.pasteCheckboxes, self.pasteColors)
        elif pasteCheckVar == 0:
            self.pasteSelectedIndices.remove(index)
            self.enableCheckboxes(self.pasteCheckboxes)
        pasteSelectedValues = [self.pasteCheckboxesText[i] for i in self.pasteSelectedIndices]
        self.pasteSelectedIndex.clear()
        for item in pasteSelectedValues:
            self.pasteSelectedIndex.append(self.excelColumnIndex(item))

    def enableCopyColumnsCheckBtn(self):
        targetRefCol = self.targetReferenceColumnCombo.get()
        if not targetRefCol == 'Select Reference Column':
            targetExcelFileName = basename(self.targetExcelFile)
            self.writeLog('info', f'[{self.targetExcelSheetCombo.get()}] is selected as target work sheet of the '
                                  f'excel file [{targetExcelFileName}]')
            self.writeLog('info', f'[{targetRefCol}] is selected as target reference column from the excel work '
                                  f'sheet [{self.targetExcelSheetCombo.get()}]')
            self.targetExcelSheetCombo.config(state='disabled', cursor='arrow')
            self.targetReferenceColumnCombo.config(state='disabled', cursor='arrow')
            self.targetReferenceColumnComboValues.remove(targetRefCol)
            for index, columnsText in enumerate(self.targetReferenceColumnComboValues):
                var = IntVar()
                self.pasteCheckBoxVars.append(var)
                pasteCheckbox = Checkbutton(self.targetPasteColumnsScrolledText, variable=var, text=columnsText,
                                            bg='light grey', state='disabled', cursor="arrow",
                                            command=lambda e=index: self.pasteCheckboxClicked(e))
                pasteCheckbox.pack()
                self.pasteCheckboxesText.append(columnsText)
                self.pasteCheckboxes[pasteCheckbox] = var
                self.targetPasteColumnsScrolledText.window_create('end', window=pasteCheckbox)
                self.targetPasteColumnsScrolledText.config(bd=4, width=12, bg='light grey', height=5)
                var.trace_add('write', self.enablePasteCheckBtnSubmitBtn)
            for copyCheckbox in self.copyCheckboxes:
                copyCheckbox.config(state='normal')
                copyCheckbox.configure(cursor="hand2")

    def enablePasteCheckBtnSubmitBtn(self, *args):
        copyCheckboxesState = [var.get() for var in self.copyCheckBoxVars]
        if any(copyCheckboxesState):
            for pasteCheckbox in self.pasteCheckboxes:
                pasteCheckbox.config(state='normal')
                pasteCheckbox.configure(cursor="hand2")
        else:
            for index, pasteCheckbox in enumerate(self.pasteCheckboxes):
                if self.pasteCheckBoxVars[index].get() == 1:
                    pasteCheckbox.deselect()
                    self.pasteSelectedIndices.clear()
                pasteCheckbox.config(state='disabled')
                pasteCheckbox.configure(cursor="arrow")
        copyCheckboxSelected = sum(copyVar.get() for copyVar in self.copyCheckBoxVars)
        pasteCheckboxSelected = sum(pasteVar.get() for pasteVar in self.pasteCheckBoxVars)
        if copyCheckboxSelected > 0 and copyCheckboxSelected == pasteCheckboxSelected:
            self.submitBtn.config(state='normal', bg='green', cursor='hand2')
        else:
            self.submitBtn.config(state='disabled', bg='light grey', cursor='arrow')

    def matchCopyPaste(self):
        def innerMatchThread():
            self.messageLabel.config(text='')
            for checkbox in self.copyCheckboxes.keys():
                checkbox.config(state='disabled')
            self.submitBtn.config(state='disabled', bg='light grey', cursor='arrow')
            self.resetBtn.config(state='disabled', bg='light grey', cursor='arrow')
            self.writeLog('info', f'Column {self.copySelectedValues} index {self.copySelectedIndex} chosen to be '
                                  f'copied')
            self.writeLog('info', f'Column {self.pasteSelectedValues} index {self.pasteSelectedIndex} chosen to be '
                                  f'pasted')
            self.writeLog('info', 'Searching... Copy & Paste')
            self.messageLabel.config(text="Processing...")
            startTime = time()
            filterwarnings("ignore", category=DeprecationWarning)
            sourceWorkbook = load_workbook(self.sourceExcelFile)
            sourceExcelSheet = self.sourceExcelSheetCombo.get()
            sourceWorkSheet = sourceWorkbook[sourceExcelSheet]
            sourceReferenceColumn = self.excelColumnIndex(self.sourceReferenceColumnCombo.get())
            sourceTotalRows = sourceWorkSheet.max_row
            self.writeLog('info', f'{sourceTotalRows} are the total unfiltered source excel sheet rows')
            targetWorkbook = load_workbook(self.targetExcelFile)
            targetExcelSheet = self.targetExcelSheetCombo.get()
            targetWorkSheet = targetWorkbook[targetExcelSheet]
            targetReferenceColumn = self.excelColumnIndex(self.targetReferenceColumnCombo.get())
            targetTotalRows = targetWorkSheet.max_row
            self.writeLog('info', f'{targetTotalRows} are the total unfiltered target excel sheet rows')
            copyPasteDone = set()
            totalSources = set()
            pasteSuccess = 0
            sourceVisibleRowCount = len([row for row in range(2, sourceTotalRows + 1) if
                                         not sourceWorkSheet.row_dimensions[row].hidden])
            self.writeLog('info', f'Total filtered source rows are {sourceVisibleRowCount}')
            source_rows = sourceWorkSheet.iter_rows(min_row=2, max_row=sourceTotalRows, values_only=True)
            for sourceRowIndex, sourceRow in enumerate(source_rows, start=2):
                if sourceWorkSheet.row_dimensions[sourceRowIndex].hidden:
                    self.writeLog('info', f'Source row [{sourceRowIndex}] is hidden(Filtered)')
                    continue
                sourceCellValue = sourceRow[sourceReferenceColumn - 1]
                self.writeLog('info', f'[{sourceCellValue}] is source cell value from [row_({sourceRowIndex}),'
                                      f'column_({sourceReferenceColumn})]')
                totalSources.add(sourceCellValue)
                target_rows = targetWorkSheet.iter_rows(min_row=2, max_row=targetTotalRows,
                                                        values_only=True)
                for targetRowIndex, targetRow in enumerate(target_rows, start=2):
                    if not targetWorkSheet.row_dimensions[targetRowIndex].hidden:
                        targetCellValue = targetRow[targetReferenceColumn - 1]
                        self.writeLog('info', f'{targetCellValue} is target cell value from [row_({targetRowIndex}),'
                                              f'column_({targetReferenceColumn})]')
                        self.writeLog('info', f'Comparing source cell value [{sourceCellValue}] with target cell value '
                                              f'[{targetCellValue}]')
                        if sourceCellValue == targetCellValue:
                            targetRow = list(targetRow)

                            for sourceColumn, targetColumn in zip(self.copySelectedIndex, self.pasteSelectedIndex):
                                targetRow[targetColumn - 1] = sourceRow[sourceColumn - 1]

                            for sourceColumn, targetColumn in zip(self.copySelectedIndex, self.pasteSelectedIndex):
                                targetCell = targetWorkSheet.cell(row=targetRowIndex, column=targetColumn)
                                sourceCell = sourceWorkSheet.cell(row=sourceRowIndex, column=sourceColumn)
                                targetCell.value = sourceCell.value
                            self.writeLog('info', f'Matched, cell Value Copied Successfully for [{targetCellValue}]')
                            if targetCellValue is not None:
                                copyPasteDone.add(targetCellValue)
                            pasteSuccess += 1
                            self.updateProgress(pasteSuccess, sourceVisibleRowCount)
                            break
            self.writeLog('info', f'Data for {len(copyPasteDone)} cells whose values are [{copyPasteDone}] copied')
            try:
                missedItem = [item.strip() for item in totalSources if
                              item is not None and item.strip() not in copyPasteDone]
                self.writeLog('info', f'Data for {len(missedItem)} cells whose values are [{missedItem}] failed to '
                                      f'copy because it is not exact match')
            except Exception as e:
                self.writeLog('error', 'Error occurred during getting missed items')
                self.writeLog('debug', f'{e}')

            self.writeLog('info', 'Saving file...')
            self.messageLabel.config(text='Saving File...')
            excelFileName = splitext(basename(self.targetExcelFile))[0]
            excelFileSave = f'{splitext(basename(self.targetExcelFile))[0]}_modified.xlsx'
            excelFIleCounter = 1
            while exists(excelFileSave):
                excelFileSave = f"{excelFileName}_modified_({excelFIleCounter}).xlsx"
                excelFIleCounter = excelFIleCounter + 1
            targetWorkbook.save(excelFileSave)
            self.writeLog('info', 'Modified file [{excelFileSave}] saved successfully')
            self.writeLog('info', 'Applying filter')
            self.progress['value'] = 99
            self.progressStyle.configure("Custom.Horizontal.TProgressbar", background='#90EE90', text='99%')
            self.messageLabel.config(text='Applying Filter...')
            modifiedFilePath = join(dirname((argv[0])), excelFileSave)
            filterApplied = list(copyPasteDone)
            try:
                filterProcess = Book(modifiedFilePath, App(visible=False))
                filterProcess.sheets[f'{targetExcelSheet}'].api.Range(
                    f"A1:{self.targetReferenceColumnComboValues[-1]}{filterProcess.sheets[f'{targetExcelSheet}'].range('A' + str(filterProcess.sheets[f'{targetExcelSheet}'].cells.last_cell.row)).end('up').row}").AutoFilter(
                    Field := targetReferenceColumn, Criterial := filterApplied, Operator := 7)
                filterProcess.save()
                filterProcess.close()
                self.writeLog('info', 'Filter applied successfully and file saved')
                self.progress['value'] = 100
                self.progressStyle.configure("Custom.Horizontal.TProgressbar", background='#7CFC00', text='100%')
            except Exception as error:
                self.writeLog('error', 'Error during filter, filter process failed')
                self.writeLog('debug', f'{error}')
            endTime = time()
            self.writeLog('info', 'TASK COMPLETED')
            elapsedTime = endTime - startTime
            hours, remainder = divmod(elapsedTime, 3600)
            minutes, remainder = divmod(remainder, 60)
            secondsTotal, millisecondsTotal = divmod(remainder, 1)
            self.writeLog('info', f'ELAPSED TIME TO COMPLETE THIS TASK IS {int(hours)} hours {int(minutes)} minutes '
                                  f'{int(secondsTotal)} seconds {int(millisecondsTotal * 1000)} milliseconds')
            self.resetBtnFunc()
            self.messageLabel.config(text=f'Changes saved in {excelFileSave}, check logs file for the status')

        threadMatchCopyPaste = Thread(target=innerMatchThread)
        threadMatchCopyPaste.start()

    def runGUI(self):

        def validateComboInput():
            return False

        comboTypeDisabledCommand = self.window.register(validateComboInput)
        self.sourceExcelSheetCombo.bind("<<ComboboxSelected>>", lambda event: self.enableSourceReferenceColumnCombo())
        self.sourceExcelSheetCombo['validatecommand'] = comboTypeDisabledCommand
        self.sourceReferenceColumnCombo.bind('<<ComboboxSelected>>', lambda event: self.enableTargetExcelDialogBtn())
        self.sourceReferenceColumnCombo['validatecommand'] = comboTypeDisabledCommand
        self.targetExcelSheetCombo.bind("<<ComboboxSelected>>", lambda event: self.enableTargetReferenceColumnCombo())
        self.targetExcelSheetCombo['validatecommand'] = comboTypeDisabledCommand
        self.targetReferenceColumnCombo.bind('<<ComboboxSelected>>', lambda event: self.enableCopyColumnsCheckBtn())
        self.targetReferenceColumnCombo['validatecommand'] = comboTypeDisabledCommand
        self.aboutLabel.bind("<Button-1>", lambda event: self.aboutWindow())

        self.window.mainloop()


if __name__ == '__main__':
    startExApp = ExApp()
    startExApp.runGUI()
